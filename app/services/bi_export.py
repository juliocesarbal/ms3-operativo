"""Chatbot BI (CU-16): convierte el reporte normalizado (bi_data) a un archivo
descargable real en PDF, Excel (.xlsx) o CSV.

reporte = {
  "titulo": str, "subtitulo": str,
  "kpis": [{"label","valor"}],
  "tablas": [{"titulo","columnas":[...],"filas":[[...]]}],
}

Devuelve (bytes, media_type, nombre_archivo).
"""
from __future__ import annotations

import csv
import io
import re
import unicodedata
from datetime import datetime
from typing import Any

MEDIA = {
    "PDF": "application/pdf",
    "EXCEL": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "CSV": "text/csv",
}
EXT = {"PDF": "pdf", "EXCEL": "xlsx", "CSV": "csv"}


def _slug(s: str) -> str:
    s = unicodedata.normalize("NFD", s or "reporte")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn").lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "reporte"


def exportar(reporte: dict[str, Any], formato: str) -> tuple[bytes, str, str]:
    formato = (formato or "PDF").upper()
    if formato == "EXCEL":
        data = _excel(reporte)
    elif formato == "CSV":
        data = _csv(reporte)
    else:
        formato = "PDF"
        data = _pdf(reporte)
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    nombre = f"{_slug(reporte.get('titulo', 'reporte'))}-{stamp}.{EXT[formato]}"
    return data, MEDIA[formato], nombre


# ---------------- CSV ----------------
def _csv(rep: dict) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([rep.get("titulo", "Reporte")])
    if rep.get("subtitulo"):
        w.writerow([rep["subtitulo"]])
    w.writerow([])
    if rep.get("kpis"):
        w.writerow(["Indicador", "Valor"])
        for k in rep["kpis"]:
            w.writerow([k["label"], k["valor"]])
        w.writerow([])
    for t in rep.get("tablas", []):
        w.writerow([t["titulo"]])
        w.writerow(t["columnas"])
        for fila in t["filas"]:
            w.writerow(fila)
        w.writerow([])
    return buf.getvalue().encode("utf-8-sig")  # BOM => acentos OK en Excel


# ---------------- Excel (.xlsx) ----------------
def _excel(rep: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="A8682F")
    head_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=15, color="1C1917")

    r = 1
    ws.cell(r, 1, rep.get("titulo", "Reporte")).font = title_font
    r += 1
    if rep.get("subtitulo"):
        ws.cell(r, 1, rep["subtitulo"]).font = Font(italic=True, color="837B70")
        r += 1
    r += 1

    if rep.get("kpis"):
        ws.cell(r, 1, "Indicadores clave").font = bold
        r += 1
        for k in rep["kpis"]:
            ws.cell(r, 1, k["label"])
            ws.cell(r, 2, k["valor"]).font = bold
            r += 1
        r += 1

    for t in rep.get("tablas", []):
        ws.cell(r, 1, t["titulo"]).font = bold
        r += 1
        for ci, col in enumerate(t["columnas"], start=1):
            c = ws.cell(r, ci, col)
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(horizontal="left")
        r += 1
        for fila in t["filas"]:
            for ci, val in enumerate(fila, start=1):
                c = ws.cell(r, ci, val)
                if isinstance(val, (int, float)):
                    c.alignment = Alignment(horizontal="right")
                    c.number_format = "#,##0"
            r += 1
        r += 1

    # ancho de columnas razonable
    for col in ws.columns:
        largo = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(largo + 2, 12), 50)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------- PDF (reportlab) ----------------
def _pdf(rep: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    accent = colors.HexColor("#A8682F")
    ink = colors.HexColor("#1C1917")
    line = colors.HexColor("#E7E1D5")
    surface2 = colors.HexColor("#FAF6EE")

    styles = getSampleStyleSheet()
    h_title = ParagraphStyle("t", parent=styles["Title"], textColor=ink, fontSize=20, spaceAfter=2)
    h_sub = ParagraphStyle("s", parent=styles["Normal"], textColor=colors.HexColor("#837B70"), fontSize=10, spaceAfter=14)
    h_sec = ParagraphStyle("h", parent=styles["Heading2"], textColor=ink, fontSize=13, spaceBefore=12, spaceAfter=6)
    foot = ParagraphStyle("f", parent=styles["Normal"], textColor=colors.HexColor("#837B70"), fontSize=8)

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=A4, topMargin=18 * mm, bottomMargin=16 * mm, leftMargin=16 * mm, rightMargin=16 * mm)
    flow: list[Any] = [Paragraph(rep.get("titulo", "Reporte"), h_title)]
    if rep.get("subtitulo"):
        flow.append(Paragraph(rep["subtitulo"], h_sub))
    if rep.get("aviso"):
        flow.append(Paragraph(f"⚠ {rep['aviso']}", h_sub))

    if rep.get("kpis"):
        flow.append(Paragraph("Indicadores clave", h_sec))
        data = [[k["label"], str(k["valor"])] for k in rep["kpis"]]
        tbl = Table(data, colWidths=[90 * mm, 80 * mm])
        tbl.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.5, line),
                    ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, surface2]),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
                    ("TEXTCOLOR", (0, 0), (-1, -1), ink),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        flow.append(tbl)

    for t in rep.get("tablas", []):
        if not t["filas"]:
            continue
        flow.append(Paragraph(t["titulo"], h_sec))
        data = [t["columnas"]] + [[str(c) for c in fila] for fila in t["filas"]]
        n = len(t["columnas"])
        ancho = (178 / n) * mm
        tbl = Table(data, colWidths=[ancho] * n, repeatRows=1)
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), accent),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, line),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, surface2]),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        flow.append(tbl)

    flow.append(Spacer(1, 14 * mm))
    flow.append(
        Paragraph(
            f"Sistema de Courier Inteligente · Grupo #11 · generado {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            foot,
        )
    )
    doc.build(flow)
    return out.getvalue()
